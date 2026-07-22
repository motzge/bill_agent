"""Excel output for bill_agent.

One workbook per month (Rechnungen_YYYY-MM.xlsx, keyed by INVOICE date),
header created on first write, rows appended below. Column pairs (Netto /
USt) exist for every plausible VAT rate so the header never has to change
after creation. Saves are atomic: write to a temp file, then replace --
a crash mid-save can never corrupt a month of existing bookings.
"""

from __future__ import annotations

import logging
import os
from datetime import datetime
from decimal import Decimal
from pathlib import Path

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from config import OUTPUT_DIR
from models import InvoiceData
from validation import PLAUSIBLE_VAT_RATES

logger = logging.getLogger(__name__)

RATES = sorted(PLAUSIBLE_VAT_RATES)  #stable column order across all files

HEADER = (
    ["Rechnungsnummer", "Rechnungsdatum", "Lieferant", "UID"]
    + [part for rate in RATES for part in (f"Netto {rate}%", f"USt {rate}%")]
    + ["Netto gesamt", "USt gesamt", "Brutto",
       "Fälligkeit", "Zahlungsziel", "Skonto", "processed_at"]
)

#extra wide columns - everything else falls back to header length
_COLUMN_WIDTHS = {
    "Rechnungsnummer": 18, "Rechnungsdatum": 14, "Lieferant": 30, "UID": 14,
    "Fälligkeit": 12, "Zahlungsziel": 38, "Skonto": 14, "processed_at": 17,
}

_DATE_FORMAT = "DD.MM.YYYY"
_MONEY_FORMAT = "#,##0.00"


def append_invoice(
    invoice: InvoiceData,
    processed_at: datetime,
    output_dir: Path = OUTPUT_DIR,
) -> bool:
    """Append one validated invoice to its monthly workbook.

    Returns False (and logs a warning) if invoice_number + supplier_name
    already exist in that workbook -- re-runs must never book twice.
    """
    path = output_dir / _month_filename(invoice)
    workbook = _load_or_create(path)
    sheet = workbook.active

    if _is_duplicate(sheet, invoice):
        logger.warning(
            "Duplicate skipped: %s / %s already in %s",
            invoice.invoice_number, invoice.supplier_name, path.name,
        )
        return False

    _append_row(sheet, invoice, processed_at)
    _save_atomic(workbook, path)
    logger.info("Booked %s into %s", invoice.invoice_number, path.name)
    return True


def _month_filename(invoice: InvoiceData) -> str:
    d = invoice.invoice_date
    return f"Rechnungen_{d.year}-{d.month:02d}.xlsx"


def _load_or_create(path: Path) -> openpyxl.Workbook:
    if path.exists():
        return openpyxl.load_workbook(path)

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Rechnungen"
    sheet.append(HEADER)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    sheet.freeze_panes = "A2"
    for index, title in enumerate(HEADER, start=1):
        width = _COLUMN_WIDTHS.get(title, max(len(title) + 2, 11))
        sheet.column_dimensions[get_column_letter(index)].width = width
    logger.info("Created new monthly workbook: %s", path.name)
    return workbook


def _dedup_key(number: str, supplier: str) -> tuple[str, str]:
    return number.strip().lower(), supplier.strip().lower()


def _is_duplicate(sheet: Worksheet, invoice: InvoiceData) -> bool:
    key = _dedup_key(invoice.invoice_number, invoice.supplier_name)
    for row in sheet.iter_rows(min_row=2, max_col=3, values_only=True):
        number, _, supplier = row
        if number is None:
            continue
        if _dedup_key(str(number), str(supplier or "")) == key:
            return True
    return False


def _append_row(sheet: Worksheet, invoice: InvoiceData, processed_at: datetime) -> None:
    by_rate = {item.rate: item for item in invoice.vat_items}
    rate_cells: list[Decimal | None] = []
    for rate in RATES:
        item = by_rate.get(rate)
        rate_cells.extend([item.net, item.tax] if item else [None, None])

    skonto = (
        f"{invoice.skonto.percent}% / {invoice.skonto.days} Tage"
        if invoice.skonto else None
    )
    row = (
        [invoice.invoice_number, invoice.invoice_date,
         invoice.supplier_name, invoice.supplier_uid]
        + rate_cells
        + [invoice.net_total, invoice.tax_total, invoice.gross_total,
           invoice.due_date, invoice.payment_term_raw, skonto, processed_at]
    )
    sheet.append(row)

    #cosmetics for the accountant: proper date and money formats
    appended = sheet[sheet.max_row]
    money_from = 4  #0-based: first rate cell after the four fixed columns
    money_to = money_from + len(RATES) * 2 + 3  #rate pairs + the three totals
    for index in range(money_from, money_to):
        appended[index].number_format = _MONEY_FORMAT
    appended[1].number_format = _DATE_FORMAT  #Rechnungsdatum
    appended[money_to].number_format = _DATE_FORMAT  #Fälligkeit
    appended[len(HEADER) - 1].number_format = "DD.MM.YYYY HH:MM"  #processed_at


def _save_atomic(workbook: openpyxl.Workbook, path: Path) -> None:
    """Write next to the target, then atomically swap it into place.

    Note for operations: if the file is open in Excel on Windows, the
    replace fails with PermissionError. main catches it, the invoice stays
    in input/ and is picked up again on the next run.
    """
    temp_path = path.with_name(path.name + ".tmp")
    workbook.save(temp_path)
    os.replace(temp_path, path)